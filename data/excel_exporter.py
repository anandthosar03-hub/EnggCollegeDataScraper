"""
Excel export functionality for college data
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List
from datetime import datetime
import os

class ExcelExporter:
    """Export college data to formatted Excel file"""
    
    @staticmethod
    def export_to_excel(data: List[dict], filename: str = None) -> str:
        """
        Export college data to Excel file
        
        Args:
            data: List of college dictionaries
            filename: Output filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"college_data_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create output directory if it doesn't exist
        output_dir = 'output'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, filename)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns
        column_order = [
            'College Name', 'University', 'Type', 'Location', 
            'Branches', 'Email', 'HOD Contact', 'Admin Contact', 
            'Other Contacts', 'Website'
        ]
        
        # Only include columns that exist
        columns = [col for col in column_order if col in df.columns]
        df = df[columns]
        
        # Export to Excel
        df.to_excel(filepath, index=False, sheet_name='Colleges')
        
        # Format the Excel file
        ExcelExporter._format_excel(filepath)
        
        return filepath
    
    @staticmethod
    def _format_excel(filepath: str):
        """
        Apply formatting to Excel file
        
        Args:
            filepath: Path to Excel file
        """
        try:
            # Load workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Set width with some padding
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save formatted workbook
            wb.save(filepath)
            
        except Exception as e:
            # If formatting fails, at least we have the data
            print(f"Warning: Could not format Excel file: {e}")
    
    @staticmethod
    def export_summary(total_colleges: int, state: str, branch: str, filepath: str) -> str:
        """
        Create a summary sheet in the Excel file
        
        Args:
            total_colleges: Number of colleges found
            state: State searched
            branch: Branch searched
            filepath: Path to Excel file
            
        Returns:
            Path to updated Excel file
        """
        try:
            wb = load_workbook(filepath)
            
            # Create summary sheet
            if 'Summary' in wb.sheetnames:
                del wb['Summary']
            
            ws_summary = wb.create_sheet('Summary', 0)
            
            # Add summary data
            summary_data = [
                ['College Scraper Report'],
                [''],
                ['Search Parameters:'],
                ['State', state],
                ['Branch', branch],
                ['Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                [''],
                ['Results:'],
                ['Total Colleges Found', total_colleges]
            ]
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            # Format summary sheet
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary['A3'].font = Font(bold=True, size=12)
            ws_summary['A8'].font = Font(bold=True, size=12)
            
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 40
            
            wb.save(filepath)
            
        except Exception as e:
            print(f"Warning: Could not create summary sheet: {e}")
        
        return filepath
